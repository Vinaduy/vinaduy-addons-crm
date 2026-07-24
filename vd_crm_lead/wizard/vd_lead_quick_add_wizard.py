# -*- coding: utf-8 -*-
"""Quick-add lead wizard — bảng tính nhập nhiều KH cùng lúc.

Mỗi dòng = 1 KH: tên, SĐT, nguồn, ngày, trạng thái.
Trạng thái rỗng → mặc định 'Khách mới'. Có chọn → lead được đẩy thẳng vào stage đó.
Auto-assign theo round-robin (leader/admin) hoặc gán cho chính mình (NV).
"""
import logging

from lxml import etree

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


SOURCE_SELECTION = [
    ('manual', '📝 Thủ công'),
    ('facebook', '📘 Facebook'),
    ('tiktok', '🎵 TikTok'),
    ('instagram', '📸 Instagram'),
    ('zalo', '💬 Zalo'),
    ('hotline', '☎ Hotline'),
    ('referral', '🤝 Giới thiệu'),
]

SOURCE_PREFIX = {
    'facebook': '[FB] ',
    'tiktok': '[TT] ',
    'instagram': '[IG] ',
    'zalo': '[Zalo] ',
    'hotline': '[Hotline] ',
    'referral': '[GT] ',
    'manual': '',
}

# Map trạng thái dashboard → stage code trong crm.stage
STATUS_SELECTION = [
    ('new', '🆕 Khách mới'),
    ('progress', '⏳ Khách đang xử lý vấn đề'),
    ('won', '🏆 Khách chốt'),
    ('lost', '❌ Khách hủy'),
]
STATUS_TO_STAGE_CODE = {
    'new': 'new',
    'progress': 'quote',  # mid-funnel representative
    'won': 'won',
    'lost': 'lost',
}


class VdLeadQuickAddWizard(models.TransientModel):
    _name = 'vd.lead.quick.add.wizard'
    _description = 'Thêm KH nhanh (bảng tính nhiều dòng)'

    line_ids = fields.One2many(
        'vd.lead.quick.add.wizard.line', 'wizard_id',
        string='Danh sách KH',
    )
    # User spec 2026-06-03: bỏ "chia đều" cũ. Flow mới: nhập KH trước → bấm
    # CHỌN NHÂN VIÊN (show_distribute=True) → chọn 1 trong 4 cách chia → cột NV
    # tự điền → bấm CHIA SỐ (action_create_leads) để tạo.
    show_distribute = fields.Boolean(default=False)
    # User spec 2026-06-05: nút CHIA SỐ chỉ hiện khi TẤT CẢ khách (Tên+SĐT) đã
    # có nhân viên. NV thường (không phải leader) tự gán mình -> luôn cho CHIA.
    can_chia_so = fields.Boolean(compute='_compute_can_chia_so')
    # User spec 2026-06-10: TẤT CẢ khách (Tên+SĐT) đã có NV → ẩn nút CHỌN NHÂN VIÊN.
    vd_all_assigned = fields.Boolean(compute='_compute_vd_all_assigned')

    @api.depends('line_ids.user_id', 'line_ids.name', 'line_ids.phone')
    def _compute_vd_all_assigned(self):
        for w in self:
            valid = w.line_ids.filtered(lambda l: l.name and l.phone)
            w.vd_all_assigned = bool(valid) and all(l.user_id for l in valid)

    @api.depends('line_ids.user_id', 'line_ids.name', 'line_ids.phone')
    def _compute_can_chia_so(self):
        is_leader = self.env.user.has_group('vd_crm_lead.vd_crm_group_team_leader')
        for w in self:
            valid = w.line_ids.filtered(lambda l: l.name and l.phone)
            if not valid:
                w.can_chia_so = False
            elif is_leader:
                w.can_chia_so = all(l.user_id for l in valid)
            else:
                w.can_chia_so = True

    # ===== NHẬP TỪ FILE EXCEL (user spec 2026-07-24) =====
    # Đẩy file .xlsx/.csv → tự lấy Tên + SĐT → nạp vào bảng → tự chia đều cho
    # TẤT CẢ nhân viên. Số nhập từ file gắn cờ vd_is_excel (thẻ FB xám đậm).
    vd_import_file = fields.Binary(string='File Excel/CSV')
    vd_import_filename = fields.Char(string='Tên file')
    # Dán danh sách 'Tên  SĐT' (mỗi dòng 1 khách) → nạp nhanh không cần file.
    vd_import_paste = fields.Text(string='Dán danh sách số')
    # Banner tóm tắt sau khi nạp file/dán (đã nạp N số, bỏ M trùng...).
    vd_import_summary = fields.Char()

    # ===== BẢNG BẬT/TẮT NV NHẬN SỐ (chuyển từ dashboard vào wizard, 2026-07-24) =====
    # Tick = NV đang nhận số. Bỏ tick + LƯU → ghi vd_can_receive_pancake=False →
    # tắt CẢ chia tự động lẫn thủ công (đồng bộ với báo cáo dashboard).
    vd_receiving_candidate_ids = fields.Many2many(
        'res.users', 'vd_qa_recv_cand_rel', 'wiz_id', 'user_id',
        compute='_compute_vd_receiving_candidates',
        string='NV có thể nhận số')
    vd_receiving_user_ids = fields.Many2many(
        'res.users', 'vd_qa_recv_on_rel', 'wiz_id', 'user_id',
        string='NV đang BẬT nhận số')

    def _vd_receiving_candidates(self):
        """POOL NV có thể nhận số (giống báo cáo dashboard): sales + trưởng nhóm/
        giám đốc, loại admin kỹ thuật."""
        Users = self.env['res.users'].sudo()
        sales = Users.search([
            ('share', '=', False), ('active', '=', True),
            ('groups_id', 'in', self.env.ref('sales_team.group_sale_salesman').id),
        ])
        bosses = Users.search([
            ('share', '=', False), ('active', '=', True),
        ]).filtered(lambda u: u.vd_crm_role in ('team_leader', 'director'))
        return (sales | bosses).filtered(
            lambda u: not (u._is_admin() or u.has_group('base.group_system'))
        ).sorted('name')

    @api.depends_context('uid')
    def _compute_vd_receiving_candidates(self):
        pool = self._vd_receiving_candidates()
        for w in self:
            w.vd_receiving_candidate_ids = [(6, 0, pool.ids)]

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        if 'vd_receiving_user_ids' in fields_list:
            pool = self._vd_receiving_candidates()
            on = pool.filtered('vd_can_receive_pancake')
            res['vd_receiving_user_ids'] = [(6, 0, on.ids)]
        return res

    def _vd_sync_receiving(self):
        """Ghi vd_can_receive_pancake cho toàn pool theo tick hiện tại — nguồn
        SỰ THẬT là bảng tick trong wizard. Đồng bộ dashboard + auto."""
        pool = self._vd_receiving_candidates()
        on_ids = set(self.vd_receiving_user_ids.ids)
        for u in pool:
            want = u.id in on_ids
            if bool(u.vd_can_receive_pancake) != want:
                u.sudo().write({'vd_can_receive_pancake': want})

    @api.onchange('vd_receiving_user_ids')
    def _onchange_vd_receiving_user_ids(self):
        """Thêm (➕) / gỡ (✕) NV ở ô tag → ghi NGAY vd_can_receive_pancake để đồng
        bộ tức thì với bảng bật/tắt bên báo cáo dashboard (user spec 2026-07-24)."""
        if self.env.user.has_group('vd_crm_lead.vd_crm_group_team_leader'):
            try:
                self._vd_sync_receiving()
            except Exception:
                pass

    def action_save_receiving(self):
        """💾 Lưu bảng BẬT/TẮT NV nhận số → ghi vd_can_receive_pancake (đồng bộ
        dashboard + chia tự động). Chia lại nếu đang ở bước chia."""
        self.ensure_one()
        self._vd_check_leader()
        self._vd_sync_receiving()
        if self.show_distribute and self.distribute_mode and self.distribute_mode != 'per_line':
            try:
                self._vd_apply_distribution()
            except UserError:
                pass
        return self._vd_reopen()

    # Gán NV hàng loạt cho các khách ĐÃ TÍCH CHỌN (user spec 2026-06-26).
    vd_bulk_user_id = fields.Many2one(
        'res.users', string='Gán NV cho khách đã chọn',
        domain="[('share', '=', False)]",
    )

    def _vd_reopen(self):
        """Mở lại wizard (giữ trạng thái) sau thao tác hàng loạt."""
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            # views tường minh → tránh crash '_preprocessAction reading map' khi
            # act_window được chạy qua doAction thiếu khoá views.
            'views': [[False, 'form']],
            'target': 'new',
            'context': dict(self.env.context, dialog_size='fullscreen'),
        }

    def action_export_excel(self):
        """Xuất danh sách khách đang nhập ra file Excel (.xlsx)."""
        self.ensure_one()
        import io
        import base64
        try:
            import xlsxwriter
        except ImportError:
            from odoo.tools.misc import xlsxwriter
        lines = self.line_ids.filtered(lambda l: l.name or l.phone)
        if not lines:
            raise UserError(_('Chưa có khách nào (Tên/SĐT) để xuất.'))
        src_map = dict(SOURCE_SELECTION)
        buf = io.BytesIO()
        wb = xlsxwriter.Workbook(buf, {'in_memory': True})
        ws = wb.add_worksheet('Khách hàng')
        hf = wb.add_format({'bold': True, 'bg_color': '#D9E1F2', 'border': 1, 'align': 'center'})
        cf = wb.add_format({'border': 1})
        headers = ['STT', 'Tên KH', 'SĐT', 'Nguồn', 'NV phụ trách',
                   'DT đất (m²)', 'Số tầng', 'Ngày']
        widths = [5, 24, 14, 14, 24, 12, 9, 14]
        for c, h in enumerate(headers):
            ws.write(0, c, h, hf)
            ws.set_column(c, c, widths[c])
        for i, l in enumerate(lines, start=1):
            ws.write(i, 0, i, cf)
            ws.write(i, 1, l.name or '', cf)
            ws.write(i, 2, l.phone or '', cf)
            ws.write(i, 3, src_map.get(l.source, l.source or ''), cf)
            ws.write(i, 4, l.user_id.name or '', cf)
            ws.write(i, 5, l.i_area_m2 or 0, cf)
            ws.write(i, 6, l.i_floors_select or '', cf)
            ws.write(i, 7, str(l.date or ''), cf)
        wb.close()
        att = self.env['ir.attachment'].create({
            'name': 'khach_hang_them_moi.xlsx',
            'type': 'binary',
            'raw': buf.getvalue(),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % att.id,
            'target': 'self',
        }

    # ==================== NHẬP TỪ FILE EXCEL / CSV ====================
    def _vd_parse_import_rows(self, data, filename):
        """Đọc bytes file .xlsx/.csv → list[list[str]] (mỗi hàng 1 list ô)."""
        import io
        fn = (filename or '').lower()
        if fn.endswith('.csv') or fn.endswith('.txt'):
            text = None
            for enc in ('utf-8-sig', 'utf-8', 'cp1258', 'latin-1'):
                try:
                    text = data.decode(enc)
                    break
                except Exception:
                    continue
            if text is None:
                text = data.decode('utf-8', 'ignore')
            import csv
            delim = ';' if text.count(';') > text.count(',') else ','
            reader = csv.reader(io.StringIO(text), delimiter=delim)
            return [[(c or '').strip() for c in r] for r in reader]
        # Mặc định: .xlsx qua openpyxl
        try:
            import openpyxl
        except ImportError:
            raise UserError(_(
                'Server chưa có thư viện openpyxl để đọc .xlsx. '
                'Hãy lưu file dạng .csv rồi đẩy lại.'))
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception:
            raise UserError(_(
                'Không mở được file — hãy chắc chắn đây là file Excel (.xlsx) '
                'hoặc CSV (.csv).'))
        ws = wb.active
        rows = []
        for r in ws.iter_rows(values_only=True):
            rows.append([('' if c is None else str(c)).strip() for c in r])
        wb.close()
        return rows

    _VD_IMP_PHONE_KEYS = ('sđt', 'sdt', 'điện thoại', 'dien thoai', 'phone',
                          'số điện', 'so dien', 'mobile', 'số đt', 'so dt', 'tel')
    _VD_IMP_NAME_KEYS = ('tên', 'ten', 'name', 'khách', 'khach', 'họ tên', 'ho ten')

    def _vd_extract_name_phone(self, rows):
        """Từ bảng thô → list[(tên, sđt)]. Ưu tiên dò cột theo tiêu đề; không có
        thì đoán theo nội dung (ô giống SĐT VN = SĐT, ô chữ dài = tên)."""
        name_col = phone_col = None
        header_idx = -1
        for i, row in enumerate(rows[:6]):
            lc = [(c or '').strip().lower() for c in row]
            pc = nc = None
            for j, c in enumerate(lc):
                if pc is None and any(k in c for k in self._VD_IMP_PHONE_KEYS):
                    pc = j
                if nc is None and any(k in c for k in self._VD_IMP_NAME_KEYS):
                    nc = j
            if pc is not None:
                phone_col, name_col, header_idx = pc, nc, i
                break
        data_rows = rows[header_idx + 1:] if header_idx >= 0 else rows
        out = []
        for row in data_rows:
            name = row[name_col] if (name_col is not None and name_col < len(row)) else ''
            phone = row[phone_col] if (phone_col is not None and phone_col < len(row)) else ''
            if not phone:
                for c in row:
                    if self._vd_phone_is_valid(c):
                        phone = c
                        break
            if not name:
                cands = [c for c in row if c and c != phone and self._vd_name_is_valid(c)]
                if cands:
                    name = max(cands, key=len)
            if phone:
                out.append(((name or '').strip(), phone.strip()))
        return out

    def action_import_excel(self):
        """Đẩy file Excel/CSV → tự lấy Tên + SĐT → nạp bảng → tự CHIA ĐỀU cho tất cả
        NV. Số nhập từ file gắn cờ Excel (thẻ Facebook xám đậm). User 2026-07-24."""
        self.ensure_one()
        self._vd_check_leader()
        if not self.vd_import_file:
            raise UserError(_('Hãy chọn file Excel/CSV trước khi bấm NHẬP FILE.'))
        import base64
        data = base64.b64decode(self.vd_import_file)
        pairs = self._vd_extract_name_phone(
            self._vd_parse_import_rows(data, self.vd_import_filename))
        if not pairs:
            raise UserError(_(
                'Không đọc được SĐT nào từ file. Kiểm tra file có cột '
                '"SĐT"/"Điện thoại"/"Phone" và có dữ liệu.'))
        self.vd_import_file = False
        self.vd_import_filename = False
        return self._vd_finish_import(pairs, source_word=_('file'))

    def _vd_parse_pasted_text(self, text):
        """Dán danh sách 'Tên  SĐT' mỗi dòng 1 khách → list[(tên, sđt)].
        Tự tách SĐT ra khỏi tên dù ở đầu/cuối, dù cách nhau bằng tab/dấu phẩy/
        khoảng trắng, dù số có khoảng trắng bên trong (0977 261 290)."""
        import re
        out = []
        for raw in (text or '').splitlines():
            line = raw.strip()
            if not line:
                continue
            m = re.search(r'(?:\+?84|0)[\s.\-]?\d[\d\s.\-]{7,13}', line)
            phone, name = '', line
            if m:
                phone = m.group(0)
                name = (line[:m.start()] + ' ' + line[m.end():])
            else:
                toks = re.split(r'[\s,;\t|]+', line)
                for i, t in enumerate(toks):
                    if self._vd_phone_is_valid(t):
                        phone = t
                        name = ' '.join(toks[:i] + toks[i + 1:])
                        break
            name = re.sub(r'[\s,;\t|]+', ' ', name).strip(' -,;|')
            if phone:
                out.append((name, phone.strip()))
        return out

    def action_import_paste(self):
        """Dán danh sách số vào ô → nạp bảng + tự CHIA ĐỀU. Thay cho đẩy file
        Excel khi chỉ cần copy vài dòng. User spec 2026-07-24."""
        self.ensure_one()
        self._vd_check_leader()
        if not (self.vd_import_paste or '').strip():
            raise UserError(_(
                'Hãy dán danh sách khách vào ô (mỗi dòng: Tên rồi SĐT).'))
        pairs = self._vd_parse_pasted_text(self.vd_import_paste)
        if not pairs:
            raise UserError(_(
                'Không tách được SĐT nào từ danh sách đã dán. Mỗi dòng cần có 1 '
                'số di động, ví dụ:\nNguyễn Ánh 0977261290'))
        self.vd_import_paste = False
        return self._vd_finish_import(pairs, source_word=_('danh sách dán'))

    def _vd_finish_import(self, pairs, source_word=''):
        """Xử lý CHUNG cho nhập-file & dán-text: lọc trùng/sai → NẠP bảng + HIỆN
        danh sách. KHÔNG tự chia — leader xem danh sách rồi bấm CHỌN NHÂN VIÊN /
        CHIA SỐ mới chia (user spec 2026-07-24). `pairs` = list[(tên, sđt-thô)]."""
        Lead = self.env['crm.lead'].sudo()

        def _core(p):
            s = Lead._vd_normalize_phones_set(p)
            return next(iter(s)) if s else ''

        seen = set()
        for l in self.line_ids:
            c = _core(l.phone)
            if c:
                seen.add(c)

        skipped_bad = skipped_dup = 0
        clean = []  # list[(name, '0xxxxxxxxx')]
        for name, phone in pairs:
            core = _core(phone)
            cphone = '0' + core if core else ''
            if not core or not self._vd_phone_is_valid(cphone):
                skipped_bad += 1
                continue
            if core in seen:
                skipped_dup += 1
                continue
            seen.add(core)
            clean.append((name, cphone))

        # Loại số ĐÃ CÓ trong hệ thống (1 truy vấn gộp, kể cả lead đã archive).
        if clean:
            variants = []
            for _n, cp in clean:
                nat = cp[1:]
                variants += [cp, nat, '84' + nat, '+84' + nat]
            existing = set()
            for r in Lead.with_context(active_test=False).search_read(
                    [('phone', 'in', variants)], ['phone']):
                cc = _core(r.get('phone'))
                if cc:
                    existing.add(cc)
            kept = []
            for name, cp in clean:
                if cp[1:] in existing:
                    skipped_dup += 1
                else:
                    kept.append((name, cp))
            clean = kept

        if not clean:
            raise UserError(_(
                'Tất cả %d số trong %s đều TRÙNG (đã có trong hệ thống) hoặc '
                'SAI định dạng — không có số mới để nhập.'
            ) % (len(pairs), source_word or _('danh sách')))

        # Nạp vào bảng: xoá các dòng trống, thêm dòng mới.
        self.line_ids.filtered(lambda l: not l.name and not l.phone).unlink()
        cmds = []
        for name, cp in clean:
            nm = self.env['crm.lead']._vd_normalize_kh_name(name) if name else ''
            if not self._vd_name_is_valid(nm):
                nm = 'Khách ' + cp[-4:]
            cmds.append((0, 0, {
                'name': nm, 'phone': cp, 'source': 'facebook',
                'vd_is_excel': True, 'status': 'new',
            }))
        self.write({'line_ids': cmds})

        # CHỈ nạp + hiện danh sách. KHÔNG tự chia số — leader xem xong rồi bấm
        # CHỌN NHÂN VIÊN → chọn cách chia → CHIA SỐ. Tóm tắt hiện ở banner.
        parts = [_('✅ Đã nạp %d số vào danh sách') % len(clean)]
        if skipped_dup:
            parts.append(_('bỏ %d trùng') % skipped_dup)
        if skipped_bad:
            parts.append(_('bỏ %d sai định dạng') % skipped_bad)
        self.vd_import_summary = (
            ' · '.join(parts)
            + _('. Kiểm tra danh sách rồi bấm CHỌN NHÂN VIÊN để chia số.'))
        return self._vd_reopen()

    def action_select_all(self):
        """Tích / BỎ tích TẤT CẢ khách (toggle)."""
        self.ensure_one()
        rows = self.line_ids.filtered(lambda l: l.name or l.phone)
        all_sel = bool(rows) and all(l.vd_selected for l in rows)
        rows.write({'vd_selected': not all_sel})
        return self._vd_reopen()

    def action_delete_selected(self):
        """XOÁ HÀNG LOẠT các khách đã tích chọn khỏi danh sách."""
        self.ensure_one()
        sel = self.line_ids.filtered('vd_selected')
        if not sel:
            raise UserError(_('Chưa tích chọn khách nào để xoá (cột ☑).'))
        sel.unlink()
        return self._vd_reopen()

    def action_assign_selected(self):
        """Gán NV (vd_bulk_user_id) cho TẤT CẢ khách đã tích chọn (vd_selected)."""
        self.ensure_one()
        if not self.vd_bulk_user_id:
            raise UserError(_('Hãy chọn nhân viên ở ô "Gán NV cho khách đã chọn" trước.'))
        sel = self.line_ids.filtered(lambda l: l.vd_selected and l.name and l.phone)
        if not sel:
            raise UserError(_('Chưa tích chọn khách nào (cột "Chọn").'))
        sel.write({'user_id': self.vd_bulk_user_id.id, 'vd_selected': False})
        self.vd_bulk_user_id = False
        return self._vd_reopen()

    distribute_mode = fields.Selection(
        [
            ('even_all', '⚖️ Chia đều cho TẤT CẢ nhân viên'),
            ('least', '📉 Dồn cho nhân viên ÍT SỐ nhất'),
            ('per_line', '✍️ Chọn nhân viên theo từng khách'),
            ('group', '👥 Chia đều theo NHÓM nhân viên đã chọn'),
        ],
        string='Cách chia số',
    )
    group_user_ids = fields.Many2many(
        'res.users', string='Nhóm NV nhận',
        domain="[('share', '=', False)]",
        help='Chọn 2+ nhân viên — khách sẽ được chia đều trong nhóm này.',
    )
    # Quick-create field: admin gõ tên → tạo vd.intake.custom.field → cột
    # tương ứng xuất hiện trong bảng (qua fields_get override trên line model).
    quick_add_field_id = fields.Many2one(
        'vd.intake.custom.field',
        string='+ Thêm cột',
        store=False,
        help='Gõ tên cột mới + Enter để tạo. Đóng và mở lại wizard để thấy cột mới.',
    )

    @api.model
    def _get_view(self, view_id=None, view_type='form', **options):
        """Xoá các <field name="extra_N"/> dư khỏi arch khi N > số custom field đang active.
        Tránh hiện 'Tuỳ chọn N' vô nghĩa trong dropdown ⋮ optional columns."""
        arch, view = super()._get_view(view_id, view_type, **options)
        try:
            n = self.env['vd.intake.custom.field'].sudo().search_count(
                [('active', '=', True)],
            )
            for idx in range(n + 1, 11):
                for node in arch.xpath(f'//field[@name="extra_{idx}"]'):
                    node.getparent().remove(node)
        except Exception:
            pass
        return arch, view

    @api.onchange('quick_add_field_id')
    def _onchange_quick_add_field_id(self):
        if self.quick_add_field_id:
            field = self.quick_add_field_id
            self.quick_add_field_id = False
            return {
                'warning': {
                    'title': '✅ Đã tạo cột mới',
                    'message': f'Cột "{field.name}" đã được tạo. Đóng wizard và '
                               f'mở lại để cột xuất hiện trong bảng.',
                },
            }

    def _vd_check_leader(self):
        if not self.env.user.has_group('vd_crm_lead.vd_crm_group_team_leader'):
            raise UserError(_(
                'Chỉ Trưởng nhóm / Admin mới được chia KH cho nhân viên.'
            ))

    def _vd_eligible_users(self):
        """POOL NV nhận KH: salesman, không phải lãnh đạo, đang nhận KH mới."""
        ResUsers = self.env['res.users'].sudo()
        candidates = ResUsers.search([
            ('share', '=', False),
            ('active', '=', True),
            ('groups_id', 'in', self.env.ref('sales_team.group_sale_salesman').id),
        ])
        # ĐỒNG BỘ (user spec 2026-07-24): công tắc vd_can_receive_pancake gate CẢ
        # chia thủ công — tắt NV ở bảng nhận số = loại khỏi pool chia đều luôn.
        return candidates.filtered(
            lambda u: not u.has_group('vd_crm_lead.vd_crm_group_team_leader')
                      and u.vd_can_receive_new_leads
                      and u.vd_can_receive_pancake
        )

    def _vd_user_new_total(self, uid):
        """Tổng KH MỚI hiện tại của 1 NV (bucket KHÁCH MỚI)."""
        Lead = self.env['crm.lead'].sudo()
        return Lead.search_count(
            Lead._dashboard_new_bucket_domain([('user_id', '=', uid)]))

    @api.model
    def _vd_phone_is_valid(self, phone):
        """SĐT hợp lệ = số di động VN: 10 số (0 + 9 số), đầu số 03/05/07/08/09.
        Chặn nhập linh tinh ('2', '3434'...)."""
        s = self.env['crm.lead']._vd_normalize_phones_set(phone)
        if not s:
            return False
        nat = next(iter(s))   # đã strip 0/84
        return len(nat) == 9 and nat[0] in '35789'

    @api.model
    def _vd_name_is_valid(self, name):
        """Tên hợp lệ = có ≥2 CHỮ CÁI (chặn 'A2', '22', ký tự linh tinh)."""
        import re
        return len(re.findall(r'[^\W\d_]', name or '', re.UNICODE)) >= 2

    def _vd_validate_quick_lines(self, lines):
        """Chặn dòng nhập linh tinh (tên/SĐT không hợp lệ) — user spec 2026-06-10."""
        bad = []
        for l in lines:
            errs = []
            if not self._vd_name_is_valid(l.name):
                errs.append('TÊN không hợp lệ')
            if not self._vd_phone_is_valid(l.phone):
                errs.append('SĐT không hợp lệ')
            if errs:
                bad.append('• %s — %s: %s' % (
                    l.name or '(trống)', l.phone or '(trống)', ', '.join(errs)))
        if bad:
            raise UserError(_(
                'Có %d dòng nhập SAI — sửa lại trước khi chia số:\n%s\n\n'
                'Quy tắc: TÊN phải có ≥2 chữ cái · SĐT phải là số di động VN '
                '(10 số, đầu 03/05/07/08/09).'
            ) % (len(bad), '\n'.join(bad)))

    def action_show_distribute(self):
        """Bấm CHỌN NHÂN VIÊN → hiện bộ chọn cách chia (sau khi đã nhập KH)."""
        self.ensure_one()
        self._vd_check_leader()
        # Tick BẬT/TẮT NV là nguồn sự thật → ghi trước khi tính pool.
        self._vd_sync_receiving()
        valid_lines = self.line_ids.filtered(lambda l: l.name and l.phone)
        if not valid_lines:
            raise UserError(_('Hãy nhập ít nhất 1 khách (Tên + SĐT) trước khi chia số.'))
        # CHẶN NHẬP LINH TINH (tên/SĐT sai) trước khi sang bước chia.
        self._vd_validate_quick_lines(valid_lines)
        # CHẶN TRÙNG NGAY (user 2026-06-05): nhập 2 số giống nhau (kể cả số chưa
        # chuẩn) hoặc đã có trong hệ thống -> không cho sang bước chia.
        dups = self.line_ids.filtered(lambda l: l.name and l.phone and l.phone_is_dup)
        if dups:
            raise UserError(_(
                'Có %d số bị TRÙNG (nhập 2 lần hoặc đã có trong hệ thống) — '
                'sửa/xoá trước khi chia số:\n%s'
            ) % (len(dups), '\n'.join(
                '• %s — %s' % (l.name or '(chưa tên)', l.phone or '') for l in dups)))
        # BẮT BUỘC chọn NGUỒN (user 2026-06-05).
        no_src = self.line_ids.filtered(lambda l: l.name and l.phone and not l.source)
        if no_src:
            raise UserError(_(
                'Vui lòng chọn NGUỒN cho các khách:\n%s'
            ) % '\n'.join('• %s — %s' % (l.name or '(chưa tên)', l.phone or '') for l in no_src))
        self.show_distribute = True
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': dict(self.env.context, dialog_size='fullscreen'),
        }

    @api.onchange('distribute_mode', 'group_user_ids')
    def _onchange_distribute_mode(self):
        """Chọn cách chia → tự ĐIỀN cột Nhân viên cho từng dòng."""
        if not self.show_distribute or not self.distribute_mode:
            return
        self._vd_apply_distribution()

    def _vd_apply_distribution(self):
        """Điền user_id cho từng dòng KH theo distribute_mode."""
        lines = self.line_ids.filtered(lambda l: l.name and l.phone)
        if not lines:
            return
        mode = self.distribute_mode
        if mode == 'per_line':
            # NV chọn tay → không tự điền (giữ nguyên / xoá để chọn lại)
            return
        if mode == 'group':
            pool = list(self.group_user_ids)
            if not pool:
                return
        else:
            pool = list(self._vd_eligible_users())
        if not pool:
            raise UserError(_(
                'Không có NV nào đủ điều kiện nhận KH mới (kiểm tra quá hạn chăm sóc).'
            ))
        # Tải hiện tại (tổng KH mới) — dùng cho 'least' và để chia đều cân bằng.
        load = {u.id: self._vd_user_new_total(u.id) for u in pool}
        # CHẶN CHIA SỐ (user spec 2026-06-12): SỨC CHỨA mỗi NV = ngưỡng - đang tồn
        # KH mới CHƯA gọi. KH wizard đều là KH MỚI chưa gọi → mỗi KH ăn 1 chỗ.
        Lead = self.env['crm.lead'].sudo()
        threshold = Lead._vd_distribute_block_threshold()
        uncalled = Lead._vd_uncalled_new_count_map([u.id for u in pool]) if threshold > 0 else {}
        remain = {
            u.id: (10 ** 9 if threshold <= 0
                   else max(0, threshold - uncalled.get(u.id, 0)))
            for u in pool
        }
        if threshold > 0 and len(lines) > sum(remain.values()):
            raise UserError(_(
                'Các NV chỉ còn nhận thêm %d khách mới (ngưỡng %d/NV) nhưng đang '
                'chia %d khách. Gọi bớt khách cũ hoặc giảm số chia, hoặc chọn cách '
                '"Tự chọn từng KH" để chia tay cho NV còn chỗ.'
            ) % (sum(remain.values()), threshold, len(lines)))
        assigned = dict(load)
        if mode == 'least':
            # Greedy: mỗi KH → NV ít số nhất CÒN sức chứa.
            for line in lines:
                avail = [u for u in pool if remain[u.id] > 0]
                u = min(avail, key=lambda x: assigned[x.id])
                line.user_id = u.id
                assigned[u.id] += 1
                remain[u.id] -= 1
        else:
            # even_all / group: round-robin theo thứ tự ít số → nhiều, BỎ QUA NV
            # hết chỗ.
            order = sorted(pool, key=lambda u: load.get(u.id, 0))
            n = len(order)
            i = 0
            for line in lines:
                guard = 0
                while remain[order[i % n].id] <= 0 and guard < n:
                    i += 1
                    guard += 1
                u = order[i % n]
                line.user_id = u.id
                remain[u.id] -= 1
                i += 1

    def action_redistribute(self):
        """Nút 'Chia lại' — áp dụng lại distribution với cấu hình hiện tại."""
        self.ensure_one()
        self._vd_sync_receiving()
        self._vd_apply_distribution()
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': dict(self.env.context, dialog_size='fullscreen'),
        }

    def action_create_leads(self):
        """Tạo N lead từ self.line_ids — mỗi dòng 1 lead."""
        self.ensure_one()
        if self.env.user.has_group('vd_crm_lead.vd_crm_group_team_leader'):
            self._vd_sync_receiving()
        Lead = self.env['crm.lead']
        ResUsers = self.env['res.users']

        lines = self.line_ids.filtered(lambda l: l.name and l.phone)
        if not lines:
            raise UserError(_('Phải có ít nhất 1 dòng có Tên + SĐT.'))

        # CHẶN NHẬP LINH TINH (tên/SĐT sai) — user spec 2026-06-10.
        self._vd_validate_quick_lines(lines)

        # CHẶN TRÙNG SỐ — không cho tạo nếu có số đã có trong hệ thống / nhập 2 lần.
        dups = lines.filtered(lambda l: l.phone_is_dup)
        if dups:
            raise UserError(_(
                'Có %d số bị TRÙNG (đã có trong hệ thống hoặc nhập 2 lần) — '
                'vui lòng sửa/xoá trước khi chia số:\n%s'
            ) % (len(dups), '\n'.join(
                '• %s — %s' % (l.name or '(chưa tên)', l.phone or '') for l in dups)))
        # BẮT BUỘC chọn NGUỒN (user 2026-06-05).
        no_src = lines.filtered(lambda l: not l.source)
        if no_src:
            raise UserError(_(
                'Vui lòng chọn NGUỒN cho các khách:\n%s'
            ) % '\n'.join('• %s — %s' % (l.name or '(chưa tên)', l.phone or '') for l in no_src))

        # CHẶN CHIA SỐ (user spec 2026-06-12): TỔNG dự kiến KH mới chưa gọi của mỗi
        # NV (đang tồn + số gán trong đợt) không vượt ngưỡng — áp cho dòng đã gán NV
        # cụ thể (nhất là cách "Tự chọn từng KH"). Auto modes đã tự tôn trọng.
        threshold = Lead._vd_distribute_block_threshold()
        if threshold > 0:
            from collections import defaultdict
            by_uid = defaultdict(int)
            for l in lines:
                if l.user_id:
                    by_uid[l.user_id.id] += 1
            if by_uid:
                current = Lead.sudo()._vd_uncalled_new_count_map(list(by_uid))
                over = []
                for uid, n in by_uid.items():
                    cap = max(0, threshold - current.get(uid, 0))
                    if n > cap:
                        over.append('• %s — đang tồn %d KH mới chưa gọi, chỉ nhận '
                                    'thêm %d (ngưỡng %d)' % (
                                        ResUsers.sudo().browse(uid).name,
                                        current.get(uid, 0), cap, threshold))
                if over:
                    raise UserError(_(
                        'KHÔNG chia được — các NV sau sẽ VƯỢT ngưỡng khách mới chưa '
                        'gọi:\n%s\n\nChọn NV khác cho các số dư (gọi bớt khách cũ để '
                        'mở thêm chỗ).'
                    ) % '\n'.join(over))

        user = self.env.user
        is_leader = user.has_group('vd_crm_lead.vd_crm_group_team_leader')
        Stage = self.env['crm.stage'].sudo()

        # Cache stage lookups for 3 status options
        stage_cache = {}
        for status_key, stage_code in STATUS_TO_STAGE_CODE.items():
            s = Stage.search([('code', '=', stage_code)], limit=1)
            if s:
                stage_cache[status_key] = s.id

        created = self.env['crm.lead']
        for line in lines:
            # Assignee: ưu tiên user chọn explicit, fallback round-robin/self
            if line.user_id:
                assignee = line.user_id
            elif is_leader:
                picked = ResUsers.sudo()._vd_pick_next_assignee()
                assignee = picked or user
            else:
                assignee = user

            prefix = SOURCE_PREFIX.get(line.source or 'manual', '')
            vals = {
                'name': f'{prefix}{line.name}'.strip(),
                'partner_name': line.name,
                'phone': line.phone,
                'user_id': assignee.id,
                'type': 'lead',
            }
            # KH nhập từ file Excel → cờ để thẻ hiện icon Facebook xám đậm.
            if line.vd_is_excel:
                vals['vd_from_excel'] = True
            stage_id = stage_cache.get(line.status)
            if stage_id:
                vals['stage_id'] = stage_id
            if line.date:
                vals['date_open'] = fields.Datetime.to_datetime(line.date)

            # Map intake mirror fields (i_*) → vd_intake_* trên crm.lead
            intake_vals = {}
            mirror_map = {
                'i_house_type': 'vd_intake_house_type',
                'i_foundation_type': 'vd_intake_foundation_type',
                # i_floors_select xử lý riêng phía dưới (vì có biến thể "Nt" = N tầng + tum)
                'i_area_m2': 'vd_intake_area_m2',
                'i_floor_1_m2': 'vd_intake_floor_1_m2',
                'i_floor_2_m2': 'vd_intake_floor_2_m2',
                'i_floor_3_m2': 'vd_intake_floor_3_m2',
                'i_floor_4_m2': 'vd_intake_floor_4_m2',
                'i_floor_5_m2': 'vd_intake_floor_5_m2',
                'i_floor_6_m2': 'vd_intake_floor_6_m2',
                'i_floor_7_m2': 'vd_intake_floor_7_m2',
                'i_floor_tum_m2': 'vd_intake_floor_tum_m2',
                'i_province_id': 'vd_intake_province_id',
                'i_district': 'vd_intake_district',
                'i_dimensions': 'vd_intake_dimensions',
                'i_land_type': 'vd_intake_land_type',
                'i_car_access_select': 'vd_intake_car_access_select',
                'i_budget_amount': 'vd_intake_budget_amount',
            }
            for line_fld, lead_fld in mirror_map.items():
                val = line[line_fld]
                if not val:
                    continue
                # Defensive: nếu lead field là Selection, chỉ gán khi key có
                # trong selection — tránh ValueError nếu wizard/lead lệch keys.
                # selection có thể là callable (extensible) → dùng helper.
                lead_field = Lead._fields.get(lead_fld)
                if lead_field and lead_field.type == 'selection':
                    sel = lead_field.selection
                    if callable(sel):
                        sel = sel(Lead)
                    valid_keys = {k for k, _lbl in (sel or [])}
                    if val not in valid_keys:
                        continue
                intake_vals[lead_fld] = val.id if hasattr(val, 'id') else val
            # Tách "Nt" → vd_intake_floors_select=N + vd_intake_has_tum=True
            if line.i_floors_select:
                raw = line.i_floors_select
                if raw.endswith('t'):
                    intake_vals['vd_intake_floors_select'] = raw[:-1]
                    intake_vals['vd_intake_has_tum'] = True
                else:
                    intake_vals['vd_intake_floors_select'] = raw
            # Nếu có diện tích Tum thì cũng set vd_intake_has_tum=True
            if intake_vals.get('vd_intake_floor_tum_m2'):
                intake_vals['vd_intake_has_tum'] = True
            vals.update(intake_vals)

            # vd_skip_assignment_balance: khách nhập TAY qua wizard là lựa chọn
            # explicit của NV/leader → KHÔNG reroute dù NV đang quá hạn (>threshold).
            # Chặn quá-hạn chỉ áp cho lead tự động (Pancake / chia đều), không áp
            # cho self-add. User spec 2026-05-30: "tự thêm thì luôn giữ".
            lead = Lead.with_context(
                vd_skip_reassign_check=True,
                vd_skip_assignment_balance=True,
            ).create(vals)
            created |= lead

            # Nếu line có preset phát sinh + qty → tạo vd.lead.surcharge
            if line.surcharge_preset_id and line.surcharge_qty:
                preset = line.surcharge_preset_id
                self.env['vd.lead.surcharge'].sudo().create({
                    'lead_id': lead.id,
                    'name': preset.name,
                    'quantity': line.surcharge_qty,
                    'quantity_label': (
                        f'Số lượng {int(line.surcharge_qty)} {preset.unit_label}'
                        if preset.unit_label else f'Số lượng {int(line.surcharge_qty)}'
                    ),
                    'unit_price': preset.unit_price,
                })

        # 🔔 THÔNG BÁO + CHUÔNG cho NV vừa được ĐẨY SỐ (user spec 2026-07-24).
        # Gửi qua bus → tab dashboard của NV kêu chuông to + hiện thông báo.
        # Không ping chính người đang đẩy (NV tự thêm mình).
        self._vd_notify_pushed(created, exclude_uid=user.id)

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Đã tạo %s KH') % len(created),
                'message': _('Đã phân bổ cho NV phụ trách.'),
                'type': 'success',
                'sticky': False,
                'next': {'type': 'ir.actions.client', 'tag': 'reload'},
            },
        }

    def _vd_notify_pushed(self, leads, exclude_uid=None):
        """Bắn thông báo realtime tới từng NV vừa được đẩy số → dashboard kêu
        chuông + hiện thông báo. Bọc try/except để KHÔNG chặn tạo lead."""
        try:
            from collections import Counter
            counts = Counter(
                l.user_id.id for l in leads
                if l.user_id and l.user_id.id != exclude_uid)
            if not counts:
                return
            Bus = self.env['bus.bus']
            pusher = self.env.user.name or ''
            ResUsers = self.env['res.users'].sudo()
            for uid, cnt in counts.items():
                partner = ResUsers.browse(uid).partner_id
                if not partner:
                    continue
                msg = {'count': cnt, 'from': pusher}
                # Odoo 18: res.partner._bus_send; fallback bus.bus._sendone.
                if hasattr(partner, '_bus_send'):
                    partner._bus_send('vd.leads.pushed', msg)
                else:
                    Bus._sendone(partner, 'vd.leads.pushed', msg)
        except Exception:
            _logger.exception('vd.leads.pushed: không gửi được thông báo bus')


class VdLeadQuickAddWizardLine(models.TransientModel):
    _name = 'vd.lead.quick.add.wizard.line'
    _description = 'Dòng KH trong wizard thêm nhanh'
    _order = 'sequence, id'

    # Base selection cho các field "extensible" — selection callable sẽ merge
    # với records từ vd.field.option để user có thể "+ Thêm mới" trên UI.
    _VD_EXT_SELECTIONS = {
        'source': SOURCE_SELECTION,
        'i_house_type': [
            ('mai_bang', 'Nhà mái bằng'),
            ('mai_thai', 'Nhà mái thái'),
            ('mai_nhat', 'Nhà mái nhật'),
            ('mai_ton', 'Nhà mái tôn'),
        ],
        'i_foundation_type': [
            ('don', 'Móng đơn'),
            ('bang', 'Móng băng'),
            ('coc', 'Móng cọc'),
        ],
        'i_dimensions': [
            ('co_so_can_phep', 'CÓ SỔ - Phải làm cấp phép'),
            ('co_so_khong_phep', 'CÓ SỔ - Không cần cấp phép'),
            ('khong_so_khong_phep', 'KHÔNG SỔ - Không cần cấp phép'),
        ],
        'i_land_type': [
            ('dat_cung', 'ĐẤT CỨNG - Liền thổ'),
            ('dat_yeu', 'ĐẤT YẾU - Ao ruộng san lấp'),
        ],
    }

    @api.model
    def _vd_ext_selection(self, fname):
        """Merge base selection + user-added records từ vd.field.option."""
        base = list(self._VD_EXT_SELECTIONS.get(fname, []))
        try:
            extras = self.env['vd.field.option'].sudo().get_options(
                self._name, fname,
            )
        except Exception:
            extras = []
        keys = {k for k, _ in base}
        return base + [(k, l) for k, l in extras if k not in keys]

    wizard_id = fields.Many2one(
        'vd.lead.quick.add.wizard', required=True, ondelete='cascade',
    )
    sequence = fields.Integer(string='STT', default=10)
    # Tích chọn nhiều khách để gán NV hàng loạt (user spec 2026-06-26).
    vd_selected = fields.Boolean(string='Chọn', default=False)
    # Dòng này được nạp từ FILE EXCEL → khi tạo lead sẽ set vd_from_excel=True
    # (thẻ hiện icon Facebook xám đậm). User spec 2026-07-24.
    vd_is_excel = fields.Boolean(default=False)
    name = fields.Char(string='Tên KH', required=False)
    phone = fields.Char(string='SĐT', required=False)

    @api.onchange('name')
    def _onchange_name_normalize(self):
        """User spec 2026-05-29: normalize tên KH theo quy tắc title-case
        (preserves ASCII team code, capitalize từng từ Vietnamese)."""
        if self.name:
            normalized = self.env['crm.lead']._vd_normalize_kh_name(self.name)
            if normalized != self.name:
                self.name = normalized
    source = fields.Selection(
        selection=lambda self: self._vd_ext_selection('source'),
        string='Nguồn',
        # User spec 2026-05-29: KHÔNG default → bắt NV chọn explicit
    )
    date = fields.Date(
        string='Ngày',
        default=fields.Date.context_today,
    )
    status = fields.Selection(
        STATUS_SELECTION,
        string='Trạng thái',
        default='new',
        help='Khách mới → stage "Khách mới". Đang xử lý → stage "Khách báo giá". '
             'Khách chốt → stage "Khách chốt".',
    )
    user_id = fields.Many2one(
        'res.users',
        string='Nhân viên',
        domain="[('share', '=', False)]",
        help='Chọn NV phụ trách. Để trống = round-robin (leader) hoặc gán cho chính mình (NV).',
    )

    # ===== Chỉ số tải NV — hiện CẠNH tên NV sau khi chia số =====
    assignee_new_total = fields.Integer(
        string='Tổng KH mới', compute='_compute_assignee_stats')
    assignee_not_called = fields.Integer(
        string='Chưa gọi', compute='_compute_assignee_stats')
    assignee_overloaded = fields.Boolean(compute='_compute_assignee_stats')
    assignee_stat_label = fields.Char(
        string='Tải NV (KH mới)', compute='_compute_assignee_stats')

    @api.depends('user_id', 'wizard_id.line_ids.user_id')
    def _compute_assignee_stats(self):
        Lead = self.env['crm.lead'].sudo()
        stat = {}

        def _stat(uid):
            if uid not in stat:
                base = Lead._dashboard_new_bucket_domain([('user_id', '=', uid)])
                stat[uid] = (Lead.search_count(base),
                             Lead.search_count(base + [('call_count', '=', 0)]))
            return stat[uid]

        # Tải hiệu dụng theo từng wizard (tổng KH mới + KH vừa gán trong đợt)
        wiz_loads = {}
        for wiz in self.mapped('wizard_id'):
            assigned = {}
            for l in wiz.line_ids:
                if l.user_id and l.name:
                    assigned[l.user_id.id] = assigned.get(l.user_id.id, 0) + 1
            loads = {uid: _stat(uid)[0] + cnt for uid, cnt in assigned.items()}
            avg = (sum(loads.values()) / len(loads)) if loads else 0
            wiz_loads[wiz.id] = (loads, avg)
        for rec in self:
            if not rec.user_id:
                rec.assignee_new_total = 0
                rec.assignee_not_called = 0
                rec.assignee_overloaded = False
                rec.assignee_stat_label = ''
                continue
            total, nc = _stat(rec.user_id.id)
            rec.assignee_new_total = total
            rec.assignee_not_called = nc
            loads, avg = wiz_loads.get(rec.wizard_id.id, ({}, 0))
            eff = loads.get(rec.user_id.id, total)
            over = bool(avg and len(loads) >= 2 and eff > avg * 1.5 and eff >= 10)
            rec.assignee_overloaded = over
            rec.assignee_stat_label = (
                '📋 %d mới · 📵 %d chưa gọi%s'
                % (total, nc, ' ⚠️ QUÁ TẢI' if over else '')
            )

    # Anchor cho widget "sao chép NV xuống" (client-side, vd_copy_user_down.js) —
    # giá trị không dùng, chỉ để re-render nút khi user_id đổi (user spec 2026-06-10).
    vd_copydown = fields.Boolean(compute='_compute_vd_copydown', store=False)

    @api.depends('user_id')
    def _compute_vd_copydown(self):
        for rec in self:
            rec.vd_copydown = bool(rec.user_id)

    # Anchor cho widget copy-down cột NGUỒN (user spec 2026-06-10).
    vd_copydown_src = fields.Boolean(compute='_compute_vd_copydown_src', store=False)

    @api.depends('source')
    def _compute_vd_copydown_src(self):
        for rec in self:
            rec.vd_copydown_src = bool(rec.source)

    # ===== CHẶN TRÙNG SỐ — đã có trong hệ thống / nhập 2 lần trong danh sách =====
    phone_is_dup = fields.Boolean(compute='_compute_phone_dup')
    phone_dup_label = fields.Char(string='Trùng', compute='_compute_phone_dup')

    # ===== CẢNH BÁO NGAY khi nhập (user spec 2026-06-10): sai tên/SĐT + trùng =====
    vd_warn_bad = fields.Boolean(compute='_compute_vd_warn')
    vd_warn_label = fields.Char(string='Cảnh báo', compute='_compute_vd_warn')

    @api.depends('name', 'phone', 'phone_is_dup', 'phone_dup_label')
    def _compute_vd_warn(self):
        Wiz = self.env['vd.lead.quick.add.wizard']
        for rec in self:
            msgs, bad = [], False
            if rec.name and not Wiz._vd_name_is_valid(rec.name):
                msgs.append('⛔ Tên sai (cần ≥2 chữ cái)')
                bad = True
            if rec.phone and not Wiz._vd_phone_is_valid(rec.phone):
                msgs.append('⛔ SĐT sai (di động VN 10 số)')
                bad = True
            if rec.phone_is_dup and rec.phone_dup_label:
                msgs.append(rec.phone_dup_label)
                bad = True
            rec.vd_warn_label = ' · '.join(msgs)
            rec.vd_warn_bad = bad

    @api.depends('phone', 'wizard_id.line_ids.phone')
    def _compute_phone_dup(self):
        Lead = self.env['crm.lead'].sudo()

        def norm(p):
            s = Lead._vd_normalize_phones_set(p)
            if s:
                return next(iter(s))
            # Fallback: số CHƯA chuẩn (vd '33') vẫn so THÔ để bắt trùng trong bảng.
            return (p or '').strip()

        # Đếm trùng TRONG wizard (cùng số nhập nhiều dòng)
        wiz_count = {}
        for wiz in self.mapped('wizard_id'):
            cnt = {}
            for l in wiz.line_ids:
                c = norm(l.phone)
                if c:
                    cnt[c] = cnt.get(c, 0) + 1
            wiz_count[wiz.id] = cnt
        # Số đã có TRONG HỆ THỐNG (mọi lead, kể cả đã archive)
        cores = {norm(l.phone) for l in self if l.phone}
        cores.discard('')
        sys_cores = set()
        if cores:
            variants = []
            for c in cores:
                variants += [c, '0' + c, '84' + c, '+84' + c]
            rows = Lead.with_context(active_test=False).search_read(
                [('phone', 'in', variants)], ['phone'])
            for r in rows:
                cc = norm(r.get('phone'))
                if cc:
                    sys_cores.add(cc)
        for rec in self:
            c = norm(rec.phone)
            if not c:
                rec.phone_is_dup = False
                rec.phone_dup_label = ''
                continue
            in_sys = c in sys_cores
            in_wiz = wiz_count.get(rec.wizard_id.id, {}).get(c, 0) > 1
            rec.phone_is_dup = bool(in_sys or in_wiz)
            if in_sys:
                rec.phone_dup_label = '⛔ TRÙNG — số đã có trong hệ thống'
            elif in_wiz:
                rec.phone_dup_label = '⛔ TRÙNG — nhập 2 lần trong danh sách'
            else:
                rec.phone_dup_label = ''

    # ===== MIRROR các trường intake từ crm.lead — admin bật tắt qua ⋮ menu =====
    # Khi NV nhập giá trị → action_create_leads sẽ ghi xuống crm.lead tương ứng.
    i_house_type = fields.Selection(
        selection=lambda self: self._vd_ext_selection('i_house_type'),
        string='Kiểu nhà',
    )
    i_foundation_type = fields.Selection(
        selection=lambda self: self._vd_ext_selection('i_foundation_type'),
        string='Loại móng',
    )
    i_floors_select = fields.Selection([
        ('1', '1 tầng'), ('1t', '1 tầng + tum'),
        ('2', '2 tầng'), ('2t', '2 tầng + tum'),
        ('3', '3 tầng'), ('3t', '3 tầng + tum'),
        ('4', '4 tầng'), ('4t', '4 tầng + tum'),
        ('5', '5 tầng'), ('5t', '5 tầng + tum'),
        ('6', '6 tầng'), ('6t', '6 tầng + tum'),
        ('7', '7 tầng'), ('7t', '7 tầng + tum'),
    ], string='Số tầng')
    i_area_m2 = fields.Float(string='Tổng DT đất (m²)', digits=(10, 1))
    # Integer (m² số nguyên) — đồng bộ với crm.lead, tránh '0,0' + parse lỗi.
    i_floor_1_m2 = fields.Float(digits=(10, 1), string='Tầng 1 (m²)')
    i_floor_2_m2 = fields.Float(digits=(10, 1), string='Tầng 2 (m²)')
    i_floor_3_m2 = fields.Float(digits=(10, 1), string='Tầng 3 (m²)')
    i_floor_4_m2 = fields.Float(digits=(10, 1), string='Tầng 4 (m²)')
    i_floor_5_m2 = fields.Float(digits=(10, 1), string='Tầng 5 (m²)')
    i_floor_6_m2 = fields.Float(digits=(10, 1), string='Tầng 6 (m²)')
    i_floor_7_m2 = fields.Float(digits=(10, 1), string='Tầng 7 (m²)')
    i_floor_tum_m2 = fields.Float(digits=(10, 1), string='Tum (m²)')
    i_province_id = fields.Many2one('res.country.state', string='Tỉnh/Thành',
                                    domain="[('country_id.code', '=', 'VN')]")
    i_district = fields.Many2one('vd.district', string='Phường/Xã')
    # Base keys PHẢI khớp 1-1 với crm.lead.vd_intake_dimensions để mirror_map
    # copy thẳng được mà không cần convert key.
    i_dimensions = fields.Selection(
        selection=lambda self: self._vd_ext_selection('i_dimensions'),
        string='Sổ đỏ / cấp phép',
    )
    i_land_type = fields.Selection(
        selection=lambda self: self._vd_ext_selection('i_land_type'),
        string='Loại đất',
    )
    i_car_access_select = fields.Selection([
        ('duoc', 'Ô tô vào được'),
        ('khong', 'Ô tô KHÔNG vào được'),
    ], string='Ô tô vào')
    i_budget_amount = fields.Float(string='Ngân sách (VNĐ)')

    # 10 cột tuỳ chọn — admin tự đặt tên qua "+ Thêm cột" (vd.intake.custom.field).
    # Label hiển thị được override dynamic trong fields_get() dựa trên config.
    extra_1 = fields.Char(string='Tuỳ chọn 1')
    extra_2 = fields.Char(string='Tuỳ chọn 2')
    extra_3 = fields.Char(string='Tuỳ chọn 3')
    extra_4 = fields.Char(string='Tuỳ chọn 4')
    extra_5 = fields.Char(string='Tuỳ chọn 5')
    extra_6 = fields.Char(string='Tuỳ chọn 6')
    extra_7 = fields.Char(string='Tuỳ chọn 7')
    extra_8 = fields.Char(string='Tuỳ chọn 8')
    extra_9 = fields.Char(string='Tuỳ chọn 9')
    extra_10 = fields.Char(string='Tuỳ chọn 10')

    # ===== Phát sinh preset (cũ — vẫn giữ phòng khi cần) =====
    surcharge_preset_id = fields.Many2one(
        'vd.lead.surcharge.preset',
        string='+ Tùy biến',
        domain="[('active', '=', True)]",
    )
    surcharge_qty = fields.Float(string='Số lượng PS', digits=(10, 2))

    @api.model
    def fields_get(self, allfields=None, attributes=None):
        """Map vd.intake.custom.field config → extra_N column labels.
        Field luôn tồn tại để view không crash; slot dư bị xoá khỏi arch trong _get_view."""
        res = super().fields_get(allfields, attributes)
        try:
            cfs = self.env['vd.intake.custom.field'].sudo().search(
                [('active', '=', True)], order='sequence, id', limit=10,
            )
            for idx, cf in enumerate(cfs, start=1):
                key = f'extra_{idx}'
                if key in res:
                    res[key]['string'] = cf.name
                    if cf.help_text:
                        res[key]['help'] = cf.help_text
        except Exception:
            pass
        return res
